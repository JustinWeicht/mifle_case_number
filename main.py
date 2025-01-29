import os
import sys
import json
import time
import signal
import datetime
import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException

# TODO

# Signal handler function
def signal_handler(sig, frame):
    print('Program terminated. Exiting...')
    sys.exit(0)

class ChromeWindowClosedException(Exception):
    pass

# Register the signal handler
signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)

def is_driver_active(driver):
    try:
        driver.title  # Accessing a property to check if the driver is still active
        return True
    except WebDriverException:
        raise ChromeWindowClosedException("Chrome browser window was closed.")

def read_excel(input_file):
    # Read the entire Excel file
    df = pd.read_excel(input_file)

    # Preprocess column names to strip spaces and make them lowercase
    df.columns = df.columns.str.strip().str.lower()

    # Define the required columns based on the new headers
    required_columns = [
        'diary code', 'atty #', 'secy #', 'coll #', 'venue', 'queue', 'diary date', 'time', 'pri',
        'file #', 'court case #', 'creditor', 'debtor', 'comment'
    ]

    # Check for missing columns
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing columns in the input file: {missing_columns}")

    # Select the required columns
    df = df[required_columns]

    return df

def get_login_creds():
    creds_path = r'C:\Users\Justin\Documents\Programming\Projects\mifile_creds\login.json'

    with open(creds_path, 'r') as file:
        creds = json.load(file)
        return creds['email'], creds['password']

def get_case_number(email, password, fileno, date):
    no_record = None
    is_efiled = None
    filing_error = None
    case_number_exists = None
    case_number = None

    # Format the date
    days_after = (datetime.datetime.strptime(date, '%m/%d/%Y') + datetime.timedelta(days=3)).strftime('%m/%d/%Y')
    days_prior = (datetime.datetime.strptime(date, '%m/%d/%Y') - datetime.timedelta(days=45)).strftime('%m/%d/%Y')

    # Initialize the WebDriver
    try:
        driver = webdriver.Chrome()
    except WebDriverException as e:
        print(f"Failed to initialize WebDriver: {str(e)}")
        return None, None, None, None, None

    try:
        # Open the website
        driver.get('https://mifile.courts.michigan.gov/')

        # Check for the modal window and close it if it appears
        try:
            modal_close_button = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'OK')]"))
            )
            modal_close_button.click()
        except TimeoutException:
            pass

        # Log in to the website
        email_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div/div[2]/form/div/div[1]/div[1]/input'))
        )
        email_field.clear()
        email_field.send_keys(email)

        password_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div/div[2]/form/div/div[1]/div[2]/div/input'))
        )
        password_field.clear()
        password_field.send_keys(password)

        # Check if the driver is still active
        if not is_driver_active(driver):
            raise WebDriverException("Chrome browser window was closed.")

        time.sleep(1)

        login_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/main/div/div/div[2]/form/div/div[2]/button"))
        )
        login_button.click()

        if not is_driver_active(driver):
            raise WebDriverException("Chrome browser window was closed.")

        # Wait for the login process to complete
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/nav/div/div[1]/div[2]/a[3]"))
        )

        # Click the History link
        history_link = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/nav/div/div[1]/div[2]/a[3]"))
        )
        history_link.click()

        # Check for the modal window and close it if it appears
        try:
            modal_close_button = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'OK')]"))
            )
            modal_close_button.click()
        except TimeoutException:
            pass

        # Ensure the History page is loaded and the button is present
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/main/div/div/filing-history/div[1]/button[2]"))
        )

        attempt = 0
        max_attempts = 5
        while attempt < max_attempts:
            try:
                network_history_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "/html/body/main/div/div/filing-history/div[1]/button[2]"))
                )
                network_history_button.click()

                if not is_driver_active(driver):
                    raise WebDriverException("Chrome browser window was closed.")

                start_date_field = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div/filing-history/div[2]/filing-history-pane/div[1]/div/div[1]/div[2]/input'))
                )
                start_date_field.clear()
                start_date_field.send_keys(days_after)

                end_date_field = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div/filing-history/div[2]/filing-history-pane/div[1]/div/div[2]/div[2]/input'))
                )
                end_date_field.clear()
                end_date_field.send_keys(days_prior)

                search_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "/html/body/main/div/div/filing-history/div[2]/filing-history-pane/div[1]/div/div[3]/div/button"))
                )
                search_button.click()

                time.sleep(1)

                # Wait for the search input field to be clickable
                search_input = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div/filing-history/div[2]/filing-history-pane/div[2]/div/filing-history-filings/div/div/div[2]/label/input'))
                )
                search_input.clear()
                search_input.send_keys(fileno)

                time.sleep(5)

                # Check for the presence of "No matching records found" element
                no_records_element_xpath = "//td[@class='dataTables_empty' and contains(text(), 'No matching records found')]"
                no_records_element_present = driver.find_elements(By.XPATH, no_records_element_xpath)
                print(f'no_records element: {no_records_element_present}')
                if no_records_element_present:
                    print(f"No matching records found for {fileno}. Proceed with e-filing.")
                    no_record = True
                    return no_record, is_efiled, filing_error, case_number_exists, case_number

                # Check for the presence of "Rejected" element
                filed_element_xpath = "/html/body/main/div/div/filing-history/div[2]/filing-history-pane/div[2]/div/filing-history-filings/div/div/table/tbody/tr[1]"
                filed_element = driver.find_element(By.XPATH, filed_element_xpath)
                print(f'filed element: {filed_element.text}')

                # Find all tr elements in the table
                table_xpath = "/html/body/main/div/div/filing-history/div[2]/filing-history-pane/div[2]/div/filing-history-filings/div/div/table/tbody/tr"
                tr_elements = driver.find_elements(By.XPATH, table_xpath)

                # Put the text from the first 2 tr_elements into a list
                tr_elements_text = [tr.text for tr in tr_elements[:2]]
                print(f'tr_elements_text: {tr_elements_text}')

                # Check for keywords in the extracted text
                if any("Payment" in text or "PAYMENT" in text or "Rejected" in text or "Refunded" in text for text in tr_elements_text):
                    print("Payment Rejected found.")
                    filing_error = True
                    return no_record, is_efiled, filing_error, case_number_exists, case_number

                elif any("Filed" in text or "Paid" in text for text in tr_elements_text):
                    print("Filed/Paid found.")
                    is_efiled = True

                # Output the result
                print(f'is_efiled: {is_efiled}')
                print(f'filing_error: {filing_error}')

                if no_records_element_present:
                    print(f"No matching records found for {fileno}. Proceed with e-filing.")
                    is_efiled = False
                    return no_record, is_efiled, filing_error, case_number_exists, case_number

                # Break the loop if the case is e-filed
                if is_efiled:
                    break

            except TimeoutException:
                print(f"Attempt {attempt + 1} of {max_attempts} failed. Retrying...")
                driver.refresh()
                attempt += 1

        if is_efiled:
            try:
                # Locate the third <td> element that contains the case number
                case_number_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[3]/span"))
                )
                case_number_text = case_number_element.get_attribute("title")
                print(f"Case Number Text: {case_number_text}")

                # Check if the case number contains "TEMP"
                if "TEMP" in case_number_text:
                    print("TEMP found in case number, ending search.")
                    case_number_exists = False
                else:
                    case_number = case_number_text
                    print(f"Case Number: {case_number}")
                    case_number_exists = True

                return no_record, True, filing_error, case_number_exists, case_number

            except TimeoutException:
                print("Timeout while trying to locate the case number element.")
                return None, None, None, None, None

    except ChromeWindowClosedException as e:
        print(f"ChromeWindowClosedException: {str(e)}")
        return None, None, None, None, None  # Return None values to indicate an error/timeout
    except WebDriverException as e:
        if not is_driver_active(driver):
            print(f"WebDriverException: {str(e)}")
            return no_record, is_efiled, filing_error, case_number_exists, case_number
        else:
            print(f"WebDriverException: {str(e)}")
            return None, None, None, None, None
    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")
        traceback.print_exc()
        return None, None, None, None, None
    finally:
        driver.quit()  # Ensure the driver is closed

# Order of operations
def main(input_file):
    # Get the login credentials
    email, password = get_login_creds()

    # Read the Excel file, load the workbook, and select the active sheet
    df = read_excel(input_file)
    workbook = load_workbook(input_file)
    sheet = workbook.active

    # Ensure the workbook and sheet are correctly loaded
    if workbook is None:
        raise ValueError("Failed to load the workbook. Please check the input file path.")
    if sheet is None:
        raise ValueError("Failed to access the active sheet. Please check the workbook.")

    # Define the fill for coloring cells
    grey_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid") # no_record fill color
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid") # filing_error is True
    blue_fill = PatternFill(start_color="FFADD9E6", end_color="FFADD9E6", fill_type="solid") # is_efiled is True - Efiled Successfully
    green_fill = PatternFill(start_color="FF98FA98", end_color="FF98FA98", fill_type="solid") # Extracted case number successfully

    # Ensure that every row is processed until no default or orange rows are left
    rows_processed = True
    while rows_processed:
        rows_processed = False

        # Process each row in the DataFrame
        for index, row in df.iterrows():
            row_index = index + 2  # +2 for 0-based index and header row

            # Check the fill color of the first cell in the row
            first_cell = sheet.cell(row=row_index, column=1)
            fill_color = first_cell.fill.start_color.rgb
            print(f"Row {row_index}: Fill Color: {fill_color}")

            # Skip the row if it has been processed (blue, green, red, or grey fill color)
            if fill_color in ["FFADD9E6", "FF98FA98", "FFFF0000", "FFD3D3D3"]:
                continue

            rows_processed = True  # Set the flag to True if any row is processed

            # Define the range of columns to color (e.g., columns A to N)
            columns_to_color = range(1, 15)  # Columns A to N (1 to 14)

            case_number_col_value = row['court case #']
            if not (pd.isna(case_number_col_value) or case_number_col_value in ["NaN", "nan", None]):
                for col_num in columns_to_color:
                    cell = sheet.cell(row=row_index, column=col_num)
                    cell.fill = green_fill
                workbook.save(input_file)
                continue

            fileno = str(int(row['file #']))
            fileno = fileno[:6]

            date = str(row['diary date'])
            date = date[:10]
            date = datetime.datetime.strptime(date, '%m/%d/%Y').strftime('%m/%d/%Y')

            no_record, is_efiled, filing_error, case_number_exists, case_number = get_case_number(email, password, fileno, date)

            if no_record is None and is_efiled is None and filing_error is None and case_number_exists is None and case_number is None:
                continue

            if no_record:
                for col_num in columns_to_color:
                    cell = sheet.cell(row=row_index, column=col_num)
                    cell.fill = grey_fill
                workbook.save(input_file)

            elif filing_error:
                for col_num in columns_to_color:
                    cell = sheet.cell(row=row_index, column=col_num)
                    cell.fill = red_fill
                workbook.save(input_file)

            elif case_number_exists and case_number is not None:
                for col_num in columns_to_color:
                    cell = sheet.cell(row=row_index, column=col_num)
                    cell.fill = green_fill
                court_case_cell = sheet.cell(row=row_index, column=11)
                court_case_cell.value = case_number
                workbook.save(input_file)

            elif is_efiled:
                for col_num in columns_to_color:
                    cell = sheet.cell(row=row_index, column=col_num)
                    cell.fill = blue_fill
                workbook.save(input_file)

    print("All rows have been processed.")

if __name__ == '__main__':
    # Test file
    input_file = r'C:\Users\Justin\Documents\Programming\Projects\output_test\mifile_case_number\Redacted S&C Diary - Copy.xlsx'

    main(input_file)
